import csv
from openpyxl import load_workbook

# Path to the Excel file
excel_file = r'C:\Users\USER\Desktop\trial ai\Imo_State_Ward_Projects.xlsx'
output_csv = r'C:\Users\USER\Desktop\trial ai\Imo_State_Ward_Projects_Consolidated.csv'

# Load the Excel workbook
wb = load_workbook(excel_file)

# Create a list to store all data
all_data = []

# Add header row
header = ['LGA', 'S/N', 'WARD', 'PROJECT TO BE EXECUTED', 'LOCATION']
all_data.append(header)

total_rows = 0
sheets_processed = 0

# Process each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Skip if sheet is empty
    if ws.max_row <= 1:
        continue
    
    sheets_processed += 1
    
    # Process each row (skip header row)
    for row_num in range(2, ws.max_row + 1):
        row_data = []
        
        # Add LGA name as first column
        row_data.append(sheet_name)
        
        # Add data from each column
        for col_num in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value is None:
                cell_value = ""
            row_data.append(str(cell_value))
        
        all_data.append(row_data)
        total_rows += 1

# Write to CSV file
with open(output_csv, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerows(all_data)

print(f"Successfully converted Excel file to CSV: {output_csv}")
print(f"Total data rows: {total_rows}")
print(f"Total LGAs processed: {sheets_processed}")

# Show first few rows
print("\nFirst 5 rows of the consolidated data:")
for i, row in enumerate(all_data[:6]):
    print(f"Row {i}: {row}")

wb.close()
